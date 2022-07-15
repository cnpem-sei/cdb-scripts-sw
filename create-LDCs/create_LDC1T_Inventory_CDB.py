from ast import Break, Pass

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import getpass

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB database
login = ItemRestApi(user, password, server, port, protocol)
print('Login successfully.\n\nReading worksheet...\n')

#Opens a Excel sheet to receive the inventory items data and save as a workbook
wb = openpyxl.load_workbook('Checklist1T.xlsx', keep_vba=True, data_only=True)
ldc_informations = wb['Principal']

#Reads the Excel and get the useful data
items_list = []
for position in range(5,100) :
    if ldc_informations.cell(row = position, column = 2 ) == None:
        pass
    item_number = ldc_informations.cell(position, 2).value
    print(item_number)
    variant = ldc_informations.cell(position, 5).value
    if '1T' == variant :
        name = 'LDC-0.3-1T:1.0:'+ item_number
        ldc_transductor = '0.3A'
        items_list.append([name, item_number, ldc_transductor])
    if '1T - 1A' == variant : 
        name = 'LDC-1.0-1T:1.0:' + item_number
        ldc_transductor = '1A'
        items_list.append([name, item_number, ldc_transductor])

#Sends information to database
for quantity in range(len(items_list)):
    #LDC transducer range information
    if items_list[quantity][2] == '1A':
        catalogid = '4212'
    if items_list[quantity][2] == '0.3A':
        catalogid = '4211'
    #Check if item already exists. If not, adds in database
    try :
        login.getItemByUniqueAttributes('Inventory', items_list[quantity][0], itemIdentifier1 = items_list[quantity][1], derivedFromItemId = catalogid)
        print('LDC %s exists.') % (items_list[quantity][1])
    except cdb.common.exceptions.objectNotFound.ObjectNotFound :
        login.addItem('Inventory', items_list[quantity][0], 'Sample', itemIdentifier1 = items_list[quantity][1], derivedFromItemId = catalogid)
        print('Upload in progress: %s %%') % (((float(quantity + 1))/len(items_list))*100)
print('Successfully created!')